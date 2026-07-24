#!/usr/bin/env python3
"""
Proof of Performance - report builder (cloud-agnostic).

Reads one or more "long" CSV files produced by the per-cloud export scripts and
builds a single Excel workbook that shows how vCPUs are distributed across CPU
vendors/architectures (Intel / AMD / ARM) over the last 12 months - month by
month AND quarter by quarter, at the account level and broken down by region.

Input CSV schema (header row required, one row per group):

    cloud,year,month,region,arch,family,vcpu_hours[,vcpus]

  cloud       e.g. AWS | Azure | GCP
  year        4-digit year (int)
  month       1-12 (int)
  region      provider region short code (e.g. us-east-1, eastus, us-central1)
  arch        Intel | AMD | ARM
  family      instance family (e.g. m7a, c7g, Dasv5, n2d) - detail only
  vcpu_hours  SUM(instance-hours x vCPUs-per-instance) for that group (float)
  vcpus       OPTIONAL provisioned vCPUs = SUM over distinct instances of each
              instance's vCPUs (count x size). A real headcount, not a
              time-average. Omit the column and the Provisioned vCPUs report
              columns are simply not emitted.

Two metrics are reported for every bucket:
  * vCPU-hours        = the raw consumption number from the CSV.
  * Provisioned vCPUs = the real count of vCPUs that existed that month (from the
                        optional vcpus column). Quarterly sheets show the PEAK
                        month, since a snapshot count can't be summed over months.

Usage:
    python3 build_report.py aws_pop_long.csv [gcp_pop_long.csv ...] -o pop_report.xlsx
    python3 build_report.py *_pop_long.csv          # combine every cloud
"""

import argparse
import calendar
import csv
import glob
import sys
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit(
        "ERROR: openpyxl is required.\n"
        "Install it with:  pip install --user openpyxl\n"
        "(In Cloud Shell this works without admin rights.)"
    )

ARCHS = ["Intel", "AMD", "ARM"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
NUM_FMT = "#,##0"
PCT_FMT = "0.0%"


def hours_in_month(year, month):
    return calendar.monthrange(year, month)[1] * 24


def quarter_of(month):
    return (month - 1) // 3 + 1


# --------------------------------------------------------------------------- #
# Load
# --------------------------------------------------------------------------- #
def load_records(files):
    records = []
    for pattern in files:
        matched = glob.glob(pattern)
        if not matched:
            matched = [pattern]  # let the open() call raise a clear error
        for path in matched:
            with open(path, newline="", encoding="utf-8-sig") as fh:
                reader = csv.DictReader(fh)
                required = {"cloud", "year", "month", "region", "arch", "vcpu_hours"}
                missing = required - set(h.strip() for h in (reader.fieldnames or []))
                if missing:
                    sys.exit(f"ERROR: {path} is missing columns: {sorted(missing)}")
                for row in reader:
                    try:
                        vh = float(row.get("vcpu_hours") or 0)
                    except ValueError:
                        vh = 0.0
                    if vh <= 0:
                        continue
                    arch = (row.get("arch") or "").strip()
                    if arch not in ARCHS:
                        arch = "Intel"
                    # vcpus (provisioned count) is optional; None = not provided
                    raw_vc = row.get("vcpus")
                    vcpus = None
                    if raw_vc not in (None, ""):
                        try:
                            vcpus = float(raw_vc)
                        except ValueError:
                            vcpus = None
                    records.append(
                        {
                            "cloud": (row.get("cloud") or "Unknown").strip(),
                            "year": int(float(row["year"])),
                            "month": int(float(row["month"])),
                            "region": (row.get("region") or "unknown").strip() or "unknown",
                            "arch": arch,
                            "family": (row.get("family") or "").strip(),
                            "vcpu_hours": vh,
                            "vcpus": vcpus,
                        }
                    )
    if not records:
        sys.exit("ERROR: no usable rows found in the input CSV(s).")
    return records


# --------------------------------------------------------------------------- #
# Aggregations
# --------------------------------------------------------------------------- #
def aggregate(records):
    # keyed dicts: key -> {arch: vcpu_hours}
    acct_month = defaultdict(lambda: defaultdict(float))   # (cloud, y, m)
    reg_month = defaultdict(lambda: defaultdict(float))    # (cloud, region, y, m)
    acct_qtr = defaultdict(lambda: defaultdict(float))     # (cloud, y, q)
    reg_qtr = defaultdict(lambda: defaultdict(float))      # (cloud, region, y, q)
    family = defaultdict(float)                            # (cloud, region, y, m, arch, family)

    # Provisioned-vCPU counts (real headcount), keyed the same way. Monthly is a
    # straight sum per arch. Quarterly is the PEAK month (max), never an average -
    # a provisioned count is a snapshot, so summing months would be meaningless.
    acct_month_vc = defaultdict(lambda: defaultdict(float))
    reg_month_vc = defaultdict(lambda: defaultdict(float))
    family_vc = defaultdict(float)
    # per-arch monthly totals collected under each quarter, to take the max later
    acct_qtr_vc_months = defaultdict(lambda: defaultdict(dict))  # (c,y,q)->arch->{(y,m):count}
    reg_qtr_vc_months = defaultdict(lambda: defaultdict(dict))
    # any vcpus data present at all? controls whether the columns are emitted
    has_vcpus = False

    # track which (year, month) actually appear, per quarter grouping, so the
    # quarterly "Avg vCPUs" denominator only counts months we have data for.
    acct_qtr_months = defaultdict(set)   # (cloud, y, q) -> {(y, m)}
    reg_qtr_months = defaultdict(set)    # (cloud, region, y, q) -> {(y, m)}

    for r in records:
        c, y, m, reg, arch, fam, vh, vc = (
            r["cloud"], r["year"], r["month"], r["region"],
            r["arch"], r["family"], r["vcpu_hours"], r["vcpus"],
        )
        q = quarter_of(m)
        acct_month[(c, y, m)][arch] += vh
        reg_month[(c, reg, y, m)][arch] += vh
        acct_qtr[(c, y, q)][arch] += vh
        reg_qtr[(c, reg, y, q)][arch] += vh
        family[(c, reg, y, m, arch, fam)] += vh
        acct_qtr_months[(c, y, q)].add((y, m))
        reg_qtr_months[(c, reg, y, q)].add((y, m))

        if vc is not None:
            has_vcpus = True
            acct_month_vc[(c, y, m)][arch] += vc
            reg_month_vc[(c, reg, y, m)][arch] += vc
            family_vc[(c, reg, y, m, arch, fam)] += vc
            am = acct_qtr_vc_months[(c, y, q)][arch]
            am[(y, m)] = am.get((y, m), 0.0) + vc
            rm = reg_qtr_vc_months[(c, reg, y, q)][arch]
            rm[(y, m)] = rm.get((y, m), 0.0) + vc

    # collapse quarter month-maps to the peak month per arch
    def peak(qmap):
        out = defaultdict(lambda: defaultdict(float))
        for key, arch_months in qmap.items():
            for arch, months in arch_months.items():
                out[key][arch] = max(months.values()) if months else 0.0
        return out

    return {
        "acct_month": acct_month,
        "reg_month": reg_month,
        "acct_qtr": acct_qtr,
        "reg_qtr": reg_qtr,
        "family": family,
        "acct_qtr_months": acct_qtr_months,
        "reg_qtr_months": reg_qtr_months,
        "acct_month_vc": acct_month_vc,
        "reg_month_vc": reg_month_vc,
        "acct_qtr_vc": peak(acct_qtr_vc_months),
        "reg_qtr_vc": peak(reg_qtr_vc_months),
        "family_vc": family_vc,
        "has_vcpus": has_vcpus,
    }


# --------------------------------------------------------------------------- #
# Sheet writing helpers
# --------------------------------------------------------------------------- #
def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws):
    for col in ws.columns:
        width = 10
        for cell in col:
            if cell.value is not None:
                width = max(width, len(str(cell.value)) + 2)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 26)


def arch_metric_headers(prefix_cols, vcpus_label=None):
    """Build the shared column header list.

    vcpus_label: header suffix for the provisioned-vCPU block, e.g. 'Provisioned
    vCPUs' (monthly) or 'Peak Provisioned vCPUs' (quarterly). None omits the block.
    """
    cols = list(prefix_cols)
    for a in ARCHS:
        cols.append(f"{a} vCPU-hrs")
    cols.append("Total vCPU-hrs")
    if vcpus_label:
        for a in ARCHS:
            cols.append(f"{a} {vcpus_label}")
        cols.append(f"Total {vcpus_label}")
    cols.append("AMD % (vCPUs)" if vcpus_label else "AMD % (vCPU-hrs)")
    cols.append("ARM % (vCPUs)" if vcpus_label else "ARM % (vCPU-hrs)")
    return cols


def write_metric_row(ws, row_idx, prefix_values, arch_hours, arch_vcpus=None):
    """arch_vcpus: {arch: provisioned vCPU count}, or None to omit that block.

    When provisioned counts are present the AMD/ARM % columns are computed from
    them (the real headcount the user cares about); otherwise from vCPU-hours.
    """
    col = 1
    for v in prefix_values:
        ws.cell(row=row_idx, column=col, value=v)
        col += 1

    total_h = sum(arch_hours.get(a, 0.0) for a in ARCHS)
    # vCPU-hours block
    for a in ARCHS:
        cell = ws.cell(row=row_idx, column=col, value=round(arch_hours.get(a, 0.0)))
        cell.number_format = NUM_FMT
        col += 1
    ws.cell(row=row_idx, column=col, value=round(total_h)).number_format = NUM_FMT
    col += 1
    # provisioned vCPUs block (real count) - only if we have the data
    if arch_vcpus is not None:
        for a in ARCHS:
            cell = ws.cell(row=row_idx, column=col, value=round(arch_vcpus.get(a, 0.0)))
            cell.number_format = NUM_FMT
            col += 1
        total_vc = sum(arch_vcpus.get(a, 0.0) for a in ARCHS)
        ws.cell(row=row_idx, column=col, value=round(total_vc)).number_format = NUM_FMT
        col += 1
    # percentages - from provisioned vCPUs when available, else vCPU-hours
    if arch_vcpus is not None:
        base = sum(arch_vcpus.get(a, 0.0) for a in ARCHS)
        amd_pct = (arch_vcpus.get("AMD", 0.0) / base) if base else 0
        arm_pct = (arch_vcpus.get("ARM", 0.0) / base) if base else 0
    else:
        amd_pct = (arch_hours.get("AMD", 0.0) / total_h) if total_h else 0
        arm_pct = (arch_hours.get("ARM", 0.0) / total_h) if total_h else 0
    ws.cell(row=row_idx, column=col, value=amd_pct).number_format = PCT_FMT
    col += 1
    ws.cell(row=row_idx, column=col, value=arm_pct).number_format = PCT_FMT
    col += 1


def build_monthly_sheet(wb, title, data_by_key, prefix_cols, vc_by_key=None):
    """Generic month sheet. key = (..prefix.., year, month)."""
    ws = wb.create_sheet(title)
    label = "Provisioned vCPUs" if vc_by_key is not None else None
    headers = arch_metric_headers(prefix_cols, label)
    ws.append(headers)
    style_header(ws, len(headers))
    ws.freeze_panes = "A2"

    for key in sorted(data_by_key.keys()):
        *prefix, year, month = key
        prefix_values = list(prefix) + [year, calendar.month_abbr[month], f"Q{quarter_of(month)}"]
        # empty dict (not None) when the sheet has the block but this row lacks
        # data, so every row's columns stay aligned to the header
        arch_vc = vc_by_key.get(key, {}) if vc_by_key is not None else None
        write_metric_row(ws, ws.max_row + 1, prefix_values, data_by_key[key], arch_vc)
    autosize(ws)
    return ws


def build_quarterly_sheet(wb, title, data_by_key, months_by_key, prefix_cols, vc_by_key=None):
    """Generic quarter sheet. key = (..prefix.., year, quarter).

    Provisioned vCPUs here is the PEAK month within the quarter (a snapshot count
    can't be summed across months), so the column is labelled accordingly.
    """
    ws = wb.create_sheet(title)
    label = "Peak Provisioned vCPUs" if vc_by_key is not None else None
    headers = arch_metric_headers(prefix_cols, label)
    ws.append(headers)
    style_header(ws, len(headers))
    ws.freeze_panes = "A2"

    for key in sorted(data_by_key.keys()):
        *prefix, year, quarter = key
        prefix_values = list(prefix) + [year, f"Q{quarter}"]
        arch_vc = vc_by_key.get(key, {}) if vc_by_key is not None else None
        write_metric_row(ws, ws.max_row + 1, prefix_values, data_by_key[key], arch_vc)
    autosize(ws)
    return ws


def build_family_sheet(wb, family_data, family_vc):
    ws = wb.create_sheet("Family-Detail")
    have_vc = bool(family_vc)
    headers = ["Cloud", "Region", "Year", "Month", "Quarter", "Arch", "Family", "vCPU-hrs"]
    if have_vc:
        headers.append("Provisioned vCPUs")
    ws.append(headers)
    style_header(ws, len(headers))
    ws.freeze_panes = "A2"
    for key in sorted(family_data.keys()):
        cloud, region, year, month, arch, fam = key
        vh = family_data[key]
        row = [cloud, region, year, calendar.month_abbr[month], f"Q{quarter_of(month)}",
               arch, fam, round(vh)]
        if have_vc:
            row.append(round(family_vc.get(key, 0.0)))
        ws.append(row)
        ws.cell(row=ws.max_row, column=8).number_format = NUM_FMT
        if have_vc:
            ws.cell(row=ws.max_row, column=9).number_format = NUM_FMT
    autosize(ws)
    return ws


def build_readme_sheet(wb, has_vcpus=False):
    ws = wb.create_sheet("About", 0)
    metric_lines = [
        ("Metrics per bucket:", True),
        ("  vCPU-hrs           = SUM(instance-hours x vCPUs per instance). Consumption.", False),
    ]
    if has_vcpus:
        metric_lines += [
            ("  Provisioned vCPUs  = SUM over DISTINCT instances of each instance's vCPUs.", False),
            ("                       A real headcount of vCPUs that existed that month", False),
            ("                       (count x size), NOT a time-average. On quarterly sheets", False),
            ("                       this is the PEAK month in the quarter.", False),
            ("  AMD % / ARM %      = share of Provisioned vCPUs on that vendor.", False),
        ]
    else:
        metric_lines += [
            ("  AMD % / ARM %      = share of vCPU-hours on that vendor.", False),
            ("  (Provisioned vCPU counts are shown when the export includes instance IDs.)", False),
        ]
    lines = [
        ("Proof of Performance - vCPU migration report", True),
        ("", False),
        ("Tracks how compute vCPUs are split across CPU vendors (Intel / AMD / ARM)", False),
        ("over the last 12 months, to show migration toward AMD.", False),
        ("", False),
    ] + metric_lines + [
        ("", False),
        ("Sheets:", True),
        ("  Account-Monthly    : per cloud, one row per month.", False),
        ("  Region-Monthly     : per cloud + region, one row per month.", False),
        ("  Account-Quarterly  : per cloud, one row per quarter.", False),
        ("  Region-Quarterly   : per cloud + region, one row per quarter.", False),
        ("  Family-Detail      : per cloud/region/month/arch/family breakdown.", False),
        ("", False),
        ("Arch = ARM covers AWS Graviton, GCP Axion/Tau (ARM), Azure Ampere.", False),
        ("Reserved/committed capacity is NOT a separate bucket here - this measures", False),
        ("the vendor of the silicon, which is what 'proof of performance' cares about.", False),
    ]
    for text, bold in lines:
        ws.append([text])
        if bold:
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 90
    return ws


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(description="Build the Proof of Performance Excel report.")
    ap.add_argument("inputs", nargs="+", help="one or more *_pop_long.csv files (globs ok)")
    ap.add_argument("-o", "--output", default="pop_report.xlsx", help="output .xlsx path")
    args = ap.parse_args()

    records = load_records(args.inputs)
    agg = aggregate(records)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default sheet; we add our own

    hv = agg["has_vcpus"]
    build_readme_sheet(wb, hv)
    build_monthly_sheet(wb, "Account-Monthly", agg["acct_month"],
                        ["Cloud", "Year", "Month", "Quarter"],
                        agg["acct_month_vc"] if hv else None)
    build_monthly_sheet(wb, "Region-Monthly", agg["reg_month"],
                        ["Cloud", "Region", "Year", "Month", "Quarter"],
                        agg["reg_month_vc"] if hv else None)
    build_quarterly_sheet(wb, "Account-Quarterly", agg["acct_qtr"],
                          agg["acct_qtr_months"], ["Cloud", "Year", "Quarter"],
                          agg["acct_qtr_vc"] if hv else None)
    build_quarterly_sheet(wb, "Region-Quarterly", agg["reg_qtr"],
                          agg["reg_qtr_months"], ["Cloud", "Region", "Year", "Quarter"],
                          agg["reg_qtr_vc"] if hv else None)
    build_family_sheet(wb, agg["family"], agg["family_vc"] if hv else None)

    wb.save(args.output)
    print(f"Wrote {args.output}")
    print(f"  {len(records)} input rows across {len({r['cloud'] for r in records})} cloud(s)")
    months = sorted({(r["year"], r["month"]) for r in records})
    if months:
        print(f"  covering {months[0][0]}-{months[0][1]:02d} to {months[-1][0]}-{months[-1][1]:02d}")


if __name__ == "__main__":
    main()
